import openpyxl

path="C:/Users/Pradeep/Desktop/shanthi_python/testdata_1.xlsx"

def get_rowcount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

def get_columncount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_column

def read_data(path,sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(rownum,columnnum).value

def write_data(path,sheetname,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(rownum,columnnum).value=data
    workbook.save(path)


